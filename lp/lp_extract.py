import json
import logging
import random
import re
import sqlite3
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

from extract_html import extract_main_job, HEADERS, JOB_TYPE_LEVEL_1
from ua_true import IdentityGenerator

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

KEYWORDS = ["餐饮业", "酒店/民宿", "旅游", "家政服务", "养老服务", "美容/美发/保健", "宠物服务"]
DB_FILE = Path("urls.db")
PROGRESS_FILE = Path("extract_progress.json")
OUTPUT_FILE = Path("提取结果.xlsx")
RANDOM_WAIT_RANGE = (1, 2)
MAX_RETRIES = 3
RETRY_BASE_WAIT = 1

COOKIES = {
    'XSRF-TOKEN': 'VC8Bd04DRDa-l9nE1Di-bA',
    '__gc_id': '8cb93083f9bb448c92a56aca96cb79ec',
    '_ga': 'GA1.1.428325617.1780621259',
    '__uuid': '1780621258602.16',
    '__sessionId': '1780621258605.89',
    'Hm_lvt_a2647413544f5a04f00da7eee0d5e200': '1780621259',
    'HMACCOUNT': 'CE1DF8FB6F6C8881',
    'user_roles': '0',
    'user_photo': '5f8fa3a9dfb13a7dee343d4808u.png',
    'user_name': '%E8%BF%9E%E7%94%B7%E5%A3%AB',
    'need_bind_tel': 'false',
    'new_user': 'false',
    'c_flag': 'a662f13d5d21f0df0c7b92770c03cd58',
    'access_system': 'C',
    'fe_se': '-1780621265486',
    'inited_user': '63293a78376373bc06255ef09bf0dd55',
    'imId_0': '12b6aeb6d47e743d561bd4fe15f0b7d2',
    'imClientId_0': '12b6aeb6d47e743d3cff5346b27f6ecb',
    'acw_tc': '7b3975a517806735711788841eeb5a29dec99a51c9e62f8be548f885154ff1',
    '_ga_54YTJKWN86': 'GS2.1.s1780673818$o4$g0$t1780673818$j60$l0$h0',
    'Hm_lpvt_a2647413544f5a04f00da7eee0d5e200': '1780673819',
    '__session_seq': '94',
    '__tlg_event_seq': '365',
}

class SkipURLError(Exception):
    pass


def table_name(keyword):
    return "urls_" + keyword.replace("/", "").replace("·", "")


def log_prefix(keyword, job_id):
    return f"kw={keyword} job={job_id}"


def extract_job_id(url):
    match = re.search(r'/a/(\d+)\.shtml', url)
    return match.group(1) if match else "unknown"


def load_progress():
    if not PROGRESS_FILE.exists():
        return {"keyword_index": 0, "url_id": 0}
    with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_progress(kw_idx, url_id):
    PROGRESS_FILE.write_text(
        json.dumps({"keyword_index": kw_idx, "url_id": url_id}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def fetch_detail(url, job_id):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            headers = IdentityGenerator.generate_headers()
            response = requests.get(url, headers=headers, timeout=20,cookies=COOKIES)

            if response.status_code in (403, 500):
                logger.error("%s 返回%s，跳过", log_prefix("", job_id), response.status_code)
                raise SkipURLError

            response.raise_for_status()
            return BeautifulSoup(response.text, "html.parser")

        except SkipURLError:
            raise
        except Exception as exc:
            if attempt >= MAX_RETRIES:
                logger.error("%s 请求失败%s次，跳过：%s", log_prefix("", job_id), MAX_RETRIES, exc)
                raise SkipURLError

            wait_seconds = RETRY_BASE_WAIT * (2 ** (attempt - 1))
            logger.warning("%s 第%s次请求失败，%s秒后重试：%s", log_prefix("", job_id), attempt, wait_seconds, exc)
            time.sleep(wait_seconds)


def save_records(records):
    if not records:
        return

    cleaned_key = re.sub(r"\s+", "", "")
    cleaned_headers = [re.sub(r"\s+", "", h) for h in HEADERS]

    if OUTPUT_FILE.exists():
        wb = load_workbook(str(OUTPUT_FILE))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "提取结果"
        ws.append(list(HEADERS))

    for rec in records:
        row = [rec.get(key, "") for key in cleaned_headers]
        ws.append(row)

    tmp_path = OUTPUT_FILE.with_suffix(".tmp")
    wb.save(str(tmp_path))
    tmp_path.replace(OUTPUT_FILE)


def crawl():
    progress = load_progress()
    start_kw = progress["keyword_index"]
    start_url_id = progress["url_id"]

    if start_kw >= len(KEYWORDS):
        logger.info("所有关键词已提取完成")
        return 0

    logger.info("从断点继续：keyword_index=%s url_id=%s", start_kw, start_url_id)

    conn = sqlite3.connect(str(DB_FILE))
    total_success = 0
    total_skip = 0

    for kw_idx in range(start_kw, len(KEYWORDS)):
        keyword = KEYWORDS[kw_idx]
        tbl = table_name(keyword)
        start_id = start_url_id if kw_idx == start_kw else 0

        rows = conn.execute(f'SELECT id, url FROM "{tbl}" WHERE id > ? ORDER BY id', (start_id,)).fetchall()

        if not rows:
            logger.info("%s 无待处理URL", keyword)
            continue

        logger.info("%s 共 %s 条URL待处理", keyword, len(rows))

        for row_id, url in rows:
            job_id = extract_job_id(url)
            prefix = log_prefix(keyword, job_id)

            wait_seconds = random.uniform(*RANDOM_WAIT_RANGE)
            logger.info("%s 等待%.2f秒后开始请求", prefix, wait_seconds)
            time.sleep(wait_seconds)

            try:
                soup = fetch_detail(url, job_id)
            except SkipURLError:
                total_skip += 1
                save_progress(kw_idx, row_id)
                continue

            try:
                rec = extract_main_job(soup)
                rec["岗位类型二级"] = keyword
                if not rec["岗位类型一级"]:
                    rec["岗位类型一级"] = JOB_TYPE_LEVEL_1
                save_records([rec])
                total_success += 1
                logger.info("%s 提取成功", prefix)
            except Exception as exc:
                logger.error("%s 解析失败：%s", prefix, exc)
                total_skip += 1

            save_progress(kw_idx, row_id)

        start_url_id = 0

    conn.close()
    logger.info("完成！成功 %s 条，跳过 %s 条，文件：%s", total_success, total_skip, OUTPUT_FILE.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(crawl())
